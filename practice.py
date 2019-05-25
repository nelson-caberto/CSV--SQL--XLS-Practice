from helpers import *

#pip install openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

if __name__ == '__main__':
	
### INIT DATABASE ###

	print("INIT DATABASE")
	
	dbConnection = initializeDatabase()
	dbCursor = dbConnection.cursor()

### READ CSV's to DB ###

	csv_file = "input_0.csv"

	print(f"Reading {csv_file} into DATABASE")

	sqlSTMT = """
	INSERT INTO a (
		id,
		int3,
		char7,
		char2,
		int4a,
		float8,
		int4b)
	VALUES (?,?,?,?,?,?,?);
	"""

	csvDBEntry(dbCursor, csv_file, sqlSTMT)	
	dbConnection.commit()
		
### COMPUTE ###

	print("COMPUTATIONS")
	
	char7s = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
	char2s = ['A', 'B']
	float8s = [
		#[row, min, max],
		(7, 0, 19999),
		(8, 20000, 24999),
		(9, 25000, 30999),
		(10, 31000, 39999),
		(11, 40000, 49999),
		(12, 50000, 69999),
		(13, 70000, 99999),
		(14, 100000, 109999),
		(15, 110000, 129999),
		(16, 130000, 169999),
		(17, 170000, 209999),
		(18, 210000, sys.maxsize)
	]

	# Get a list possibilities of int4a in this dataset
	sqlSTMT = f'SELECT DISTINCT int4a FROM a;'
	try:
		dbCursor.execute(sqlSTMT)
	except sqlite3.Error as e:
		print(f'Database Error: {e} on {sqlSTMT}')
	except Exception as e:
		print(f'Exception in _query: {e} on {sqlSTMT}')
	int4as = [int4a[0] for int4a in dbCursor.fetchall()]
	int4as.sort()

	results = {}
	for int4a in int4as:
		results[int4a] = {}
		for char2 in char2s:
			results[int4a][char2] = {}
			for char7 in char7s:
				results[int4a][char2][char7] = {}
				for (row, min, max) in float8s:
					sqlSTMT = f'SELECT COUNT(*) FROM a WHERE int4a={int4a} AND char2="{char2}" AND char7="{char7}" AND float8 BETWEEN {min} AND {max}'
					try:
						dbCursor.execute(sqlSTMT)
					except sqlite3.Error as e:
						print(f"Database error: {e} on {sqlSTMT}")
						continue
					except Exception as e:
						print(f"Exception in _query: {e} on {sqlSTMT}")
						continue
					results[int4a][char2][char7][row] = dbCursor.fetchone()[0]

	totals = {}
	for int4a in int4as:
		totals[int4a] = {}
		for (row, min, max) in float8s:
			sqlSTMT = f'SELECT COUNT(*) FROM a WHERE int4a={int4a} AND float8 BETWEEN {min} and {max};'
			try:
				dbCursor.execute(sqlSTMT)
			except sqlite3.Error as e:
				print(f'Database error: {e} on {sqlSTMT}')
				continue
			except Exception as e:
				print(f'Exception in _query: {e} on {sqlSTMT}')
				continue
			totals[int4a][row] = dbCursor.fetchone()[0]

	dbCursor.close()
	
### OUTPUT ###

	xlsx_file = "output_0.xlsx"
	print(f"Writing {xlsx_file}")

	wb = Workbook()
	ws = wb.active
	offset = 0

	cell = lambda column, row: get_column_letter(column+offset)+str(row)

	for int4a in int4as:

		for column in range(1,18):
			for row in range(1,19):
				ws[cell(column, row)] = writeCell(column, row, int4a, results, totals, float8s)
				
		offset += 17

	wb.save(xlsx_file)
